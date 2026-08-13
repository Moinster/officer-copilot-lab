"""Generate the fictional practice datasets for the Kraft Heinz Executive Copilot Lab.

All output is synthetic and deterministic (fixed seed). It describes a fictitious
CPG manufacturer, "Lakeshore Foods Co.", selling to fictitious retail customers.
It is not Kraft Heinz data, financials, consumer records, or performance results.

Usage:  python tools/generate_demo_data.py
Writes: assets/Demo_Weekly_Customer_Category_Performance.xlsx
        assets/Demo_Consumer_Care_Feedback.csv
"""

from __future__ import annotations

import csv
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SEED = 20260813
ROOT = Path(__file__).resolve().parents[1]
ASSETS = ROOT / "assets"

DISCLAIMER = (
    "Fictitious demo data for Microsoft 365 Copilot training. "
    "Not real customer, employee, consumer, vendor, or financial data."
)

WEEKS = [date(2026, 5, 30) + timedelta(days=7 * i) for i in range(6)]

RED = "E4002B"
INK = "1A1110"

# --------------------------------------------------------------------------- #
# Customer / channel performance
# --------------------------------------------------------------------------- #

REGIONS = {
    "Northeast": ("NE", 1.00),
    "Southeast": ("SE", 1.05),
    "Midwest": ("MW", 0.97),
    "South Central": ("SC", 0.90),
    "West": ("WE", 1.02),
}

CUSTOMERS = [
    # name, channel, base weekly net sales, yoy bias, service bias
    ("Contoso Markets", "Grocery", 1_850_000, 0.030, 0.00),
    ("Fabrikam Club", "Club", 2_450_000, 0.062, 0.01),
    ("Northwind Grocers", "Grocery", 1_120_000, -0.008, -0.01),
    ("Adventure Works Foods", "Mass", 3_100_000, 0.041, 0.00),
    ("Tailspin Convenience", "Convenience", 640_000, 0.015, -0.02),
]

CUSTOMER_ISSUES = [
    "Case fill rate below service target",
    "On-shelf availability gap on core items",
    "Promotion executed late at store level",
    "Forecast accuracy deteriorating",
    "Trade spend running above plan",
    "Distribution loss on a core item",
]

# --------------------------------------------------------------------------- #
# Category / brand performance
# --------------------------------------------------------------------------- #

CATEGORY_TREE = [
    ("Taste Elevation", "Ketchup & Tomato", ["Ketchup Squeeze", "Ketchup Glass", "Tomato Paste"]),
    ("Taste Elevation", "Sauces & Dressings", ["Mayonnaise", "BBQ Sauce", "Salad Dressing"]),
    ("Easy Meals", "Boxed Meals", ["Boxed Macaroni", "Skillet Meal Kits"]),
    ("Easy Meals", "Canned Meals", ["Canned Pasta", "Canned Beans", "Ready Soups"]),
    ("Substantial Snacking", "Meat Snacks", ["Meat Sticks", "Jerky Multipack"]),
    ("Substantial Snacking", "Cheese Snacks", ["Snack Cheese Cups", "String Cheese"]),
    ("Coffee & Beverages", "Coffee", ["Ground Coffee", "Single Serve Coffee"]),
    ("Coffee & Beverages", "Powdered Beverages", ["Drink Mix Canister", "Drink Mix Sticks"]),
    ("Cheese & Dairy", "Cream Cheese", ["Cream Cheese Brick", "Cream Cheese Tub"]),
    ("Cheese & Dairy", "Sliced & Shredded", ["Cheese Slices", "Shredded Cheese"]),
]

CATEGORY_ISSUES = [
    "Low sell-through / inventory build",
    "Out-of-stock risk at retail",
    "Margin compression vs. prior year",
    "Promotion did not lift volume as planned",
    "Elevated days of supply",
    "Share loss to private label",
]

REGIONAL_NARRATIVE = {
    "Northeast": (
        "Convert Club momentum into full-portfolio distribution",
        "",
        "Tighten replenishment on core ketchup and sauces; secure incremental club displays",
    ),
    "Southeast": (
        "Extend snacking growth into Convenience doors",
        "",
        "Audit promo execution with the Grocery team; expand snacking distribution in Convenience",
    ),
    "Midwest": (
        "Recover Easy Meals volume with a value pack offer",
        "",
        "Run a service-recovery plan on Easy Meals and rebalance inventory across DCs",
    ),
    "South Central": (
        "Reset the weakest customer relationships in the region",
        "",
        "Priority intervention: joint forecasting reset, service recovery, and a distribution rebuild",
    ),
    "West": (
        "Scale eCommerce share on Coffee & Beverages",
        "",
        "Shift promotional dollars toward higher-return digital and club events",
    ),
}


def build_regional_summary(customer_rows: list[list]) -> list[list]:
    """Derive the summary tab from the generated rows so evidence always reconciles."""
    metrics = {
        "net sales YoY": 6,
        "on-shelf availability": 10,
        "case fill rate": 11,
        "promo compliance": 12,
        "forecast accuracy": 15,
    }

    def average(rows: list[list], index: int) -> float:
        return sum(r[index] for r in rows) / len(rows)

    company = {label: average(customer_rows, index) for label, index in metrics.items()}

    summary: list[list] = []
    for region, (opportunity, _placeholder_risk, action) in REGIONAL_NARRATIVE.items():
        rows = [r for r in customer_rows if r[1] == region]
        regional = {label: average(rows, index) for label, index in metrics.items()}
        weakest = min(regional, key=lambda label: regional[label] - company[label])
        gap = regional[weakest] - company[weakest]
        if gap < 0:
            risk = (
                f"Below company average on {weakest}: {regional[weakest]:.1%} "
                f"vs. {company[weakest]:.1%} ({gap * 100:+.1f} pts)"
            )
        else:
            strongest = max(regional, key=lambda label: regional[label] - company[label])
            risk = (
                f"No metric below company average this period; watch {weakest} "
                f"({regional[weakest]:.1%}). Strength: {strongest} at {regional[strongest]:.1%}"
            )
        evidence = "; ".join(f"{label} avg {regional[label]:+.1%}" if label == "net sales YoY"
                             else f"{label} avg {regional[label]:.1%}" for label in metrics)
        summary.append([region, opportunity, risk, action, evidence])
    return summary


ASSUMPTIONS = [
    "All data is fictional and generated programmatically for illustrative training use only.",
    "This dataset is not tied to any real company, customer, employee, consumer, vendor, or financial results.",
    "Company (fictional): Lakeshore Foods Co., a large packaged food and beverage manufacturer.",
    "Retail customers (fictional): Contoso Markets, Fabrikam Club, Northwind Grocers, Adventure Works Foods, Tailspin Convenience.",
    "Time period: six consecutive weeks ending 2026-05-30 through 2026-07-04.",
    "Currency values are in US dollars and represent net sales unless otherwise labeled.",
    "Volume is expressed in equivalent cases.",
    "Percentages are stored as decimals (0.954 = 95.4%) and formatted for display.",
    "Case Fill Rate is the share of ordered cases shipped complete and on time.",
    "On-Shelf Availability is the share of store-level audits where the item was found on shelf.",
    "Distribution Points approximate total distribution points (TDP) for the customer's carried assortment.",
    "Trade Spend % is promotional and trade investment as a share of gross sales.",
    "Forecast Accuracy is 1 minus the weighted absolute percentage error at the customer level.",
    "Key Issue is populated only when a metric crosses a training threshold; blanks are intentional.",
    "Deliberate outliers exist so that analysis exercises have something meaningful to find.",
    "Category platforms are generic descriptors and do not represent any real brand or product.",
    "No consumer, employee, or personally identifiable information is present in this dataset.",
]

# --------------------------------------------------------------------------- #
# Consumer care feedback
# --------------------------------------------------------------------------- #

FEEDBACK_TEMPLATES = [
    (
        "Product availability",
        "Negative",
        "Availability / Distribution",
        "Escalate to the customer team; review replenishment on core items",
        [
            "My store has been out of my usual size for three weeks now.",
            "The shelf tag is there but the product never is.",
            "I can only find the large size, never the one I actually buy.",
        ],
    ),
    (
        "Packaging",
        "Negative",
        "Packaging / Design",
        "Route to packaging engineering; review closure and opening force",
        [
            "The new cap is very hard to open and I needed help with it.",
            "The seal split before I got the package home from the store.",
            "The resealable tab stopped sticking after the first use.",
        ],
    ),
    (
        "Price and value",
        "Negative",
        "Pricing / Value",
        "Share with revenue management; evaluate pack-price architecture",
        [
            "The price went up and the package looks smaller than it used to.",
            "It costs more here than at the club store for the same size.",
            "Hard to justify the price when the store brand is right next to it.",
        ],
    ),
    (
        "Promotion and coupon",
        "Negative",
        "Promotion / Trade Execution",
        "Confirm promo setup with the customer team; verify coupon redemption path",
        [
            "The advertised deal did not ring up correctly at checkout.",
            "The digital coupon would not load to my loyalty account.",
            "The display sign said one price and the register said another.",
        ],
    ),
    (
        "Product experience",
        "Neutral",
        "Product Quality / Consistency",
        "Log with quality and R&D for consistency review; no safety concern indicated",
        [
            "The texture seemed different from the last package I bought.",
            "Portion size in the box seemed smaller than I remembered.",
            "Flavor was fine but not quite what I expected from the new recipe.",
        ],
    ),
    (
        "Labeling and nutrition",
        "Neutral",
        "Labeling / Transparency",
        "Route to regulatory and labeling; assess clarity of on-pack information",
        [
            "The nutrition panel is hard to read on the smaller package.",
            "I could not tell from the front of pack whether this was the reduced sodium version.",
            "Ingredient list wording changed and I was not sure why.",
        ],
    ),
    (
        "eCommerce and delivery",
        "Negative",
        "eCommerce / Fulfillment",
        "Work with the retailer's digital team on substitution rules and packing standards",
        [
            "My online order substituted a completely different flavor without asking.",
            "The delivery arrived with the outer carton crushed.",
            "The item showed available online and then was refunded after delivery.",
        ],
    ),
    (
        "Customer service",
        "Positive",
        "Consumer Care",
        "Share positive feedback with the care team; no action needed",
        [
            "The person I spoke with resolved my issue on the first call.",
            "I got a response to my email the same day, which I did not expect.",
            "Follow-up was quick and I did not have to repeat my story.",
        ],
    ),
    (
        "Product experience",
        "Positive",
        "Product Quality / Consistency",
        "Share positive feedback with the brand team; no action needed",
        [
            "The new pack size is exactly right for my household.",
            "Quality has been consistent every time I buy it.",
            "My family switched back after trying the new variety.",
        ],
    ),
    (
        "Product availability",
        "Neutral",
        "Availability / Distribution",
        "Monitor; confirm whether the gap is assortment or replenishment driven",
        [
            "Found most of what I wanted, one variety was missing.",
            "The club pack is usually there but the single is not.",
            "Availability improved compared with last month.",
        ],
    ),
]

CARE_CHANNELS = ["Consumer Care Line", "Email", "Social", "Retailer Review", "Survey", "Brand Website"]
CONSUMER_SEGMENTS = ["Household with children", "Single-person household", "Value seeker", "Convenience seeker", "Loyalist"]
RESOLUTION_STATUS = ["Open", "In Progress", "Resolved", "Escalated"]


def build_customer_rows(rng: random.Random) -> list[list]:
    rows: list[list] = []
    for region, (code, region_factor) in REGIONS.items():
        for index, (customer, channel, base_sales, yoy_bias, service_bias) in enumerate(CUSTOMERS, start=1):
            zone = f"{code}-Z{index:02d}"
            # South Central is the deliberate problem region.
            trouble = region == "South Central"
            for week in WEEKS:
                noise = rng.uniform(0.94, 1.07)
                net_sales = round(base_sales * region_factor * noise, -2)
                yoy = round(yoy_bias + rng.uniform(-0.035, 0.035) + (-0.075 if trouble else 0.0), 4)
                cases = int(net_sales / rng.uniform(23.0, 31.0))
                price_mix = round(rng.uniform(-0.012, 0.041), 4)
                avg_net_price = round(net_sales / cases, 2)
                osa = round(min(0.995, rng.uniform(0.925, 0.985) + service_bias - (0.06 if trouble else 0.0)), 4)
                fill = round(min(0.999, rng.uniform(0.930, 0.992) + service_bias - (0.055 if trouble else 0.0)), 4)
                promo_compliance = round(min(0.995, rng.uniform(0.845, 0.985) - (0.05 if trouble else 0.0)), 4)
                tdp = int(rng.uniform(1400, 5200) * region_factor)
                trade_spend = round(rng.uniform(0.148, 0.226) + (0.02 if trouble else 0.0), 4)
                forecast_accuracy = round(rng.uniform(0.780, 0.945) - (0.08 if trouble else 0.0), 4)
                escalations = rng.choices([0, 0, 0, 1, 1, 2, 3], weights=[26, 18, 14, 16, 12, 9, 5])[0]
                if trouble:
                    escalations += rng.choice([0, 1, 2])

                issue = None
                if fill < 0.93:
                    issue = CUSTOMER_ISSUES[0]
                elif osa < 0.93:
                    issue = CUSTOMER_ISSUES[1]
                elif promo_compliance < 0.87:
                    issue = CUSTOMER_ISSUES[2]
                elif forecast_accuracy < 0.78:
                    issue = CUSTOMER_ISSUES[3]
                elif trade_spend > 0.215:
                    issue = CUSTOMER_ISSUES[4]
                elif yoy < -0.06:
                    issue = CUSTOMER_ISSUES[5]

                rows.append([
                    week, region, zone, customer, channel, net_sales, yoy, cases, price_mix,
                    avg_net_price, osa, fill, promo_compliance, tdp, trade_spend,
                    forecast_accuracy, escalations, issue,
                ])
    return rows


def build_category_rows(rng: random.Random) -> list[list]:
    rows: list[list] = []
    for platform, category, brand_groups in CATEGORY_TREE:
        for brand_group in brand_groups:
            base = rng.uniform(380_000, 1_650_000)
            trend = rng.uniform(-0.055, 0.085)
            weak = brand_group in {"Canned Pasta", "Drink Mix Canister", "Boxed Macaroni"}
            for week in WEEKS:
                net_sales = round(base * rng.uniform(0.92, 1.09), -2)
                yoy = round(trend + rng.uniform(-0.03, 0.03) - (0.07 if weak else 0.0), 4)
                margin = round(rng.uniform(0.238, 0.412) - (0.045 if weak else 0.0), 4)
                cases = int(net_sales / rng.uniform(21.0, 34.0))
                days_supply = round(rng.uniform(24.0, 78.0) + (22.0 if weak else 0.0), 1)
                sell_through = round(rng.uniform(0.22, 0.61) - (0.09 if weak else 0.0), 4)
                promo = rng.choices(["Yes", "No"], weights=[35, 65])[0]
                ecom_share = round(rng.uniform(0.03, 0.19), 4)
                share = round(rng.uniform(0.14, 0.42) - (0.03 if weak else 0.0), 4)

                issue = None
                if sell_through < 0.24:
                    issue = CATEGORY_ISSUES[0]
                elif days_supply < 28:
                    issue = CATEGORY_ISSUES[1]
                elif margin < 0.24:
                    issue = CATEGORY_ISSUES[2]
                elif promo == "Yes" and yoy < 0:
                    issue = CATEGORY_ISSUES[3]
                elif days_supply > 82:
                    issue = CATEGORY_ISSUES[4]
                elif yoy < -0.07:
                    issue = CATEGORY_ISSUES[5]

                rows.append([
                    week, platform, category, brand_group, net_sales, yoy, margin, cases,
                    days_supply, sell_through, promo, ecom_share, share, issue,
                ])
    return rows


def write_rows(ws, rows: list[list], start_row: int = 3) -> None:
    for row_index, row in enumerate(rows, start=start_row):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)


def style_sheet(ws, headers: list[str], formats: dict[int, str], widths: dict[int, int]) -> None:
    ws.cell(row=1, column=1, value=DISCLAIMER).font = Font(italic=True, color="7A2140", size=9)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col, fmt in formats.items():
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = fmt
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"


def build_workbook(rng: random.Random) -> Workbook:
    wb = Workbook()
    customer_rows = build_customer_rows(rng)

    # --- Customer Performance ---
    ws = wb.active
    ws.title = "Customer Performance"
    headers = [
        "Week Ending", "Region", "Zone", "Retail Customer", "Channel", "Net Sales",
        "Net Sales YoY %", "Volume (Cases)", "Price/Mix %", "Average Net Price",
        "On Shelf Availability %", "Case Fill Rate %", "Promo Compliance %",
        "Distribution Points", "Trade Spend %", "Forecast Accuracy %",
        "Service Escalations", "Key Issue",
    ]
    write_rows(ws, customer_rows)
    style_sheet(
        ws, headers,
        formats={1: "yyyy-mm-dd", 6: '"$"#,##0', 7: "0.0%", 8: "#,##0", 9: "0.0%",
                 10: '"$"#,##0.00', 11: "0.0%", 12: "0.0%", 13: "0.0%", 14: "#,##0",
                 15: "0.0%", 16: "0.0%"},
        widths={1: 13, 2: 15, 3: 10, 4: 23, 5: 14, 6: 13, 7: 15, 8: 15, 9: 11,
                10: 16, 11: 20, 12: 16, 13: 18, 14: 18, 15: 14, 16: 18, 17: 18, 18: 34},
    )

    # --- Category Performance ---
    ws2 = wb.create_sheet("Category Performance")
    headers2 = [
        "Week Ending", "Category Platform", "Category", "Brand Group", "Net Sales",
        "Net Sales YoY %", "Gross Margin %", "Volume (Cases)", "Days of Supply",
        "Sell Through %", "Promo Flag", "eCommerce Share %", "Market Share %", "Key Issue",
    ]
    write_rows(ws2, build_category_rows(rng))
    style_sheet(
        ws2, headers2,
        formats={1: "yyyy-mm-dd", 5: '"$"#,##0', 6: "0.0%", 7: "0.0%", 8: "#,##0",
                 9: "0.0", 10: "0.0%", 12: "0.0%", 13: "0.0%"},
        widths={1: 13, 2: 22, 3: 22, 4: 22, 5: 13, 6: 15, 7: 15, 8: 15, 9: 15,
                10: 15, 11: 12, 12: 18, 13: 15, 14: 34},
    )

    # --- Regional Summary ---
    ws3 = wb.create_sheet("Regional Summary")
    headers3 = ["Region", "Top Opportunity", "Top Risk", "Recommended Action", "Evidence Metric"]
    write_rows(ws3, build_regional_summary(customer_rows))
    style_sheet(ws3, headers3, formats={}, widths={1: 16, 2: 46, 3: 46, 4: 58, 5: 46})
    for row in range(3, ws3.max_row + 1):
        for col in range(1, 6):
            ws3.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    # --- Notes and Assumptions ---
    ws4 = wb.create_sheet("Notes and Assumptions")
    ws4.cell(row=1, column=1, value=DISCLAIMER).font = Font(italic=True, color="7A2140", size=9)
    ws4.cell(row=3, column=1, value="Assumptions").font = Font(bold=True, color=RED, size=12)
    for index, note in enumerate(ASSUMPTIONS, start=4):
        ws4.cell(row=index, column=2, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 110

    return wb


def build_feedback_rows(rng: random.Random) -> list[list]:
    regions = list(REGIONS)
    customers = [c[0] for c in CUSTOMERS]
    brand_lookup = [(p, b) for p, _c, groups in CATEGORY_TREE for b in groups]

    rows: list[list] = []
    start = date(2026, 5, 30)
    for day_offset in range(0, 78):
        if rng.random() < 0.05:
            continue
        entry_date = start + timedelta(days=day_offset)
        topic, sentiment, impact, follow_up, texts = rng.choice(FEEDBACK_TEMPLATES)
        platform, brand_group = rng.choice(brand_lookup)
        region = rng.choices(regions, weights=[18, 20, 19, 26, 17])[0]
        status = rng.choices(RESOLUTION_STATUS, weights=[22, 26, 36, 16])[0]
        if sentiment == "Positive":
            status = rng.choice(["Resolved", "Resolved", "In Progress"])
        rows.append([
            entry_date.isoformat(),
            rng.choice(CARE_CHANNELS),
            region,
            rng.choice(customers),
            platform,
            brand_group,
            rng.choice(CONSUMER_SEGMENTS),
            topic,
            sentiment,
            rng.choice(texts),
            status,
            impact,
            follow_up,
        ])
    return rows


def write_feedback_csv(rows: list[list], path: Path) -> None:
    headers = [
        "Date", "Channel", "Region", "Retail Customer", "Category Platform", "Brand Group",
        "Consumer Segment", "Topic", "Sentiment", "Feedback Text", "Resolution Status",
        "Impact Area", "Recommended Follow-Up",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        handle.write(f"# {DISCLAIMER}\n")
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def main() -> None:
    ASSETS.mkdir(parents=True, exist_ok=True)
    rng = random.Random(SEED)

    workbook = build_workbook(rng)
    workbook_path = ASSETS / "Demo_Weekly_Customer_Category_Performance.xlsx"
    workbook.save(workbook_path)

    feedback_path = ASSETS / "Demo_Consumer_Care_Feedback.csv"
    write_feedback_csv(build_feedback_rows(rng), feedback_path)

    print(f"wrote {workbook_path.relative_to(ROOT)}")
    print(f"wrote {feedback_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
